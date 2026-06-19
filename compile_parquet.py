#!/usr/bin/env python3
"""
compile_parquet.py - v3.1 (handles BOTH cleaned and raw VAHAN xlsx formats)

Format A (cleaned, e.g. Haryana_3W_AllFuel_2026.xlsx):
    Row 1 : title
    Row 2 : "S. No." | "Maker" | "Jan 2026" | "Feb 2026" | ... | "Total"
    Row 3+: data

Format B (raw VAHAN, e.g. West_Bengal_4W_AllFuel_2026_20260618.xlsx):
    Row 1 : title
    Row 2 : "S" | "Maker" | (merged "Month Wise") | "TOTA"
    Row 3 : "No" | <blank>  | <blanks across>  | "L"
    Row 4 : <blank> | <blank> | "JAN" | "FEB" | "MAR" | ...
    Row 5+: data

For each unique (state, combo, year), picks the file with the LARGEST date
suffix and ignores older duplicates. Output: long-format master.parquet with
columns:
    state | vehicle_category | fuel_type | year | month | maker
        | registrations | source_file | source_date
"""

import argparse, os, re, sys, glob, logging, warnings
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------- CLI -----------------------------------------------
parser = argparse.ArgumentParser(description="Compile Vahan xlsx -> master.parquet")
parser.add_argument("--data-dir", default="./data", help="Folder with xlsx files")
parser.add_argument("--out",      default="master.parquet", help="Output parquet path")
args = parser.parse_args()

# ---------------- Logging -------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("compile_parquet")

# ---------------- Constants -----------------------------------------
MONTH_ABBR = ["JAN","FEB","MAR","APR","MAY","JUN",
              "JUL","AUG","SEP","OCT","NOV","DEC"]
MONTH_TO_NUM = {m: i+1 for i, m in enumerate(MONTH_ABBR)}

# Filename: <State>_<2W|3W|4W>_<PureEV|AllFuel>_<YYYY>[_<YYYYMMDD>].xlsx
FNAME_RE = re.compile(
    r"^(?P<state>.+?)_(?P<combo>(?:2W|3W|4W)_(?:PureEV|AllFuel))_"
    r"(?P<year>\d{4})(?:_(?P<date>\d{8}))?\.xlsx$"
)

# ---------------- Filename parser -----------------------------------
def parse_filename(fname):
    m = FNAME_RE.match(fname)
    if not m:
        return None
    d = m.groupdict()
    cat, fuel = d["combo"].split("_")
    return {
        "state":            d["state"].replace("_", " ").replace("and", "&"),
        "state_safe":       d["state"],
        "vehicle_category": cat,
        "fuel_type":        fuel,
        "combo":            d["combo"],
        "year":             int(d["year"]),
        "date_suffix":      d["date"] or "00000000",
        "filename":         fname,
    }

# ---------------- Pick latest per group ------------------------------
def pick_latest_files(data_dir):
    all_files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")))
    log.info(f"Scanning {len(all_files)} xlsx files in {data_dir}")

    parsed, skipped = [], []
    for fp in all_files:
        info = parse_filename(os.path.basename(fp))
        if info is None:
            skipped.append(fp)
            continue
        info["filepath"] = fp
        parsed.append(info)

    if skipped:
        log.warning(f"Skipped {len(skipped)} files (bad name pattern):")
        for s in skipped[:5]:
            log.warning(f"   - {os.path.basename(s)}")

    if not parsed:
        log.error("No parseable files found!")
        return [], 0

    df = pd.DataFrame(parsed)
    df_sorted = df.sort_values("date_suffix", ascending=False)
    df_latest = df_sorted.drop_duplicates(
        subset=["state_safe", "combo", "year"], keep="first"
    )
    dropped = len(df) - len(df_latest)

    log.info(f"Picked {len(df_latest)} latest files; ignored {dropped} older duplicates")
    log.info("Sample of selected files:")
    for _, row in df_latest.head(5).iterrows():
        log.info(f"   [{row['date_suffix']}] {row['filename']}")
    return df_latest.to_dict("records"), dropped

# ---------------- Unmerge + load -------------------------------------
def unmerge_and_load(filepath):
    """Load xlsx, unmerge all merged cells (propagate top-left value)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    for mr in list(ws.merged_cells.ranges):
        tlv = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                ws.cell(r, c, tlv)
    return wb, ws

def ws_to_df(ws):
    """Convert openpyxl worksheet (all values) to pandas DataFrame, no header."""
    data = list(ws.values)
    return pd.DataFrame(data)

# ---------------- Detect month tokens in a cell ----------------------
def extract_months_from_cell(val):
    """
    Return list of month-numbers (1-12) found in this single cell value.
    Handles: 'JAN', 'Jan', 'Jan 2026', 'JAN-2026', 'jan_2026', 'JAN26', etc.
    """
    if val is None:
        return []
    s = str(val).strip().upper()
    if not s:
        return []
    # Replace separators with spaces
    for ch in ("-", "/", "_", ".", ",", ":"):
        s = s.replace(ch, " ")
    tokens = s.split()
    months = []
    for tok in tokens:
        if len(tok) >= 3:
            first3 = tok[:3]
            if first3 in MONTH_TO_NUM:
                months.append(MONTH_TO_NUM[first3])
    return months

# ---------------- Locate header row + maker col ----------------------
def find_structure(df):
    """
    Scan first 10 rows. Return (header_row_idx, month_col_map, maker_col).
    month_col_map: {col_idx: month_num} ; maker_col: int
    Returns (None, {}, None) if month row not found.
    """
    scan_rows = min(10, len(df))
    best_row, best_month_map = None, {}

    # Skip row 0 (title) — start from row 1
    for ri in range(1, scan_rows):
        month_map = {}
        for ci in range(len(df.columns)):
            val = df.iat[ri, ci] if ci < df.shape[1] else None
            months_in_cell = extract_months_from_cell(val)
            if months_in_cell:
                # one cell normally yields one month; take the first
                month_map[ci] = months_in_cell[0]
        # Score = number of distinct months found in this row
        distinct = len(set(month_map.values()))
        if distinct > len(set(best_month_map.values())):
            best_month_map = month_map
            best_row = ri

    if best_row is None or len(best_month_map) < 2:
        return None, {}, None

    # Find Maker column: scan rows 1..header_row (skip row 0 title)
    # STRICT exact match only — title like "Maker Month Wise Data of..." must NOT match
    maker_col = None
    for ri in range(max(1, best_row - 3), best_row + 1):
        for ci in range(df.shape[1]):
            val = df.iat[ri, ci]
            if val is None:
                continue
            s = str(val).strip().upper()
            if s in ("MAKER", "MAKERS"):
                maker_col = ci
                break
        if maker_col is not None:
            break

    # Fallback: column 1 (B)
    if maker_col is None:
        maker_col = 1

    return best_row, best_month_map, maker_col

# ---------------- Read one file --------------------------------------
def read_one(file_info):
    fp    = file_info["filepath"]
    fname = file_info["filename"]
    try:
        wb, ws = unmerge_and_load(fp)
        df_raw = ws_to_df(ws)
        wb.close()
    except Exception as e:
        log.warning(f"   [READ FAIL] {fname}: {e}")
        return None

    if df_raw.empty:
        log.warning(f"   [EMPTY] {fname}")
        return None

    header_row, month_map, maker_col = find_structure(df_raw)

    if header_row is None or not month_map:
        log.warning(f"   [NO MONTHS] {fname}")
        return None

    n_months = len(set(month_map.values()))
    log.info(f"   [{fname}] header_row={header_row}, maker_col={maker_col}, months={n_months}")

    # Extract data rows starting from row AFTER the month header row
    data_rows = []
    sorted_month_cols = sorted(month_map.items(), key=lambda x: x[1])

    for ri in range(header_row + 1, len(df_raw)):
        row = df_raw.iloc[ri]
        # Maker
        maker_val = row.iloc[maker_col] if maker_col < len(row) else None
        if maker_val is None or (isinstance(maker_val, float) and pd.isna(maker_val)):
            continue
        maker_str = str(maker_val).strip()
        upper     = maker_str.upper()
        if not maker_str:
            continue
        if upper in ("NAN", "MAKER", "MAKERS", "S. NO.", "S.NO.", "SNO", "S NO"):
            continue
        if upper.startswith("TOTAL") or upper.startswith("GRAND TOTAL"):
            continue

        # Months
        all_zero = True
        row_recs = []
        for col_idx, month_num in sorted_month_cols:
            v = row.iloc[col_idx] if col_idx < len(row) else 0
            if isinstance(v, str):
                v = v.replace(",", "").strip()
            v_num = pd.to_numeric(v, errors="coerce")
            reg   = int(v_num) if not pd.isna(v_num) else 0
            if reg != 0:
                all_zero = False
            row_recs.append((month_num, reg))

        # Skip rows where everything is zero (phantom rows)
        if all_zero:
            continue

        for month_num, reg in row_recs:
            data_rows.append({
                "state":            file_info["state"],
                "vehicle_category": file_info["vehicle_category"],
                "fuel_type":        file_info["fuel_type"],
                "year":             file_info["year"],
                "month":            month_num,
                "maker":            maker_str,
                "registrations":    reg,
                "source_file":      fname,
                "source_date":      file_info["date_suffix"],
            })

    if not data_rows:
        log.warning(f"   [NO DATA] {fname} - header found but no maker rows")
        return None

    df_long = pd.DataFrame(data_rows)
    log.info(f"   [OK] {fname}: {len(df_long)} rows extracted")
    return df_long

# ---------------- Main ----------------------------------------------
def main():
    if not os.path.isdir(args.data_dir):
        log.error(f"Data dir not found: {args.data_dir}")
        sys.exit(1)

    latest_files, dropped = pick_latest_files(args.data_dir)
    if not latest_files:
        log.error("Nothing to compile")
        sys.exit(1)

    frames, fail_count = [], 0
    for fi in latest_files:
        df = read_one(fi)
        if df is None or df.empty:
            fail_count += 1
            continue
        frames.append(df)

    if not frames:
        log.error("All file reads returned empty - nothing written")
        sys.exit(1)

    master = pd.concat(frames, ignore_index=True)
    master.to_parquet(args.out, index=False, compression="snappy")

    size_mb = os.path.getsize(args.out) / (1024 * 1024)
    log.info("=" * 60)
    log.info("COMPILE SUMMARY")
    log.info("=" * 60)
    log.info(f"  Output            : {args.out}  ({size_mb:.2f} MB)")
    log.info(f"  Files used        : {len(latest_files)}")
    log.info(f"  Older duplicates  : {dropped}  (ignored)")
    log.info(f"  Files failed read : {fail_count}")
    log.info(f"  Total rows        : {len(master):,}")
    log.info(f"  Unique states     : {master['state'].nunique()}")
    log.info(f"  Unique makers     : {master['maker'].nunique()}")
    log.info(f"  Years covered     : {sorted(master['year'].unique().tolist())}")
    log.info(f"  Categories        : {sorted(master['vehicle_category'].unique().tolist())}")
    log.info(f"  Fuel types        : {sorted(master['fuel_type'].unique().tolist())}")
    log.info(f"  Source-date range : {master['source_date'].min()} -> {master['source_date'].max()}")
    log.info("=" * 60)

if __name__ == "__main__":
    main()
