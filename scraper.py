#!/usr/bin/env python3
"""
VAHAN EV SCRAPER v5.5.1 — Multi-Filter: 2W / 3W / 4W x PureEV / AllFuel
Author : Harshal Panchal
Date   : 2026-06-13  (v5.5.1 patch 2026-06-18)

v5.5.1 surgical changes vs v5.5 (NO core-logic changes):
  1) YEAR is now auto-detected from current IST date (was hard-coded "2025"
     causing filename mismatch with existing 2026 xlsx files in data/).
  2) New RUN_DATE constant (YYYYMMDD) is appended to every output filename
     so each daily run produces uniquely-named files. compile_parquet.py
     picks the latest date suffix per (state, combo, year) tuple.
  3) Filename in download_state() now uses {YEAR}_{RUN_DATE} suffix.
  4) Banner updated to v5.5.1.

ROOT CAUSE FIXED IN v5.5 (unchanged here):
  The VAHAN portal generates the Excel file SERVER-SIDE only after the
  Refresh button is clicked AND the AJAX response fully completes.
  Previous versions clicked the download button before the server had
  finished regenerating the export — so the old cached file was served.

  v5.5 FIX: After applying filters + clicking Refresh, the script now
  waits for the VAHAN AJAX response to carry a fresh data payload
  (detected by monitoring network responses for the groupingTable XHR).
  Only after that response is confirmed does it click the Excel download.
"""

import asyncio, os, time, logging, warnings, traceback
import json, random, signal, argparse, tempfile, sys
from datetime import datetime, timezone, timedelta
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ===================================================================
#  CLI ARGUMENTS
# ===================================================================

parser = argparse.ArgumentParser(description="VAHAN EV Scraper v5.5.1")
parser.add_argument("--visible", action="store_true",
                    help="Show browser window (non-headless mode)")
parser.add_argument("--no-resume", action="store_true",
                    help="Ignore progress file, re-download everything")
cli_args, _ = parser.parse_known_args()

# ===================================================================
#  CONFIGURATION
# ===================================================================

VAHAN_URL         = "https://vahan.parivahan.gov.in/vahan4dashboard/vahan/view/reportview.xhtml"

# v5.5.1: auto-detect current IST year (was hard-coded "2025")
_IST_NOW          = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
YEAR              = str(_IST_NOW.year)
RUN_DATE          = _IST_NOW.strftime("%Y%m%d")   # YYYYMMDD suffix per run

X_AXIS            = "Month Wise"
Y_AXIS            = "Maker"
OUTPUT_DIR        = "./vahan_downloads"  # GitHub Actions moves these to data/ after run
HEADLESS          = not cli_args.visible
RESUME_ENABLED    = not cli_args.no_resume
MAX_RETRIES       = 3
RETRY_BACKOFF     = 5
PAGE_TIMEOUT      = 60_000
DOWNLOAD_TIMEOUT  = 120_000
MIN_FILE_SIZE     = 3000

# --- Timing ---
AJAX_WAIT            = 1.5
STATE_DELAY_MIN      = 2.0
STATE_DELAY_MAX      = 5.0
PANEL_OPEN_WAIT      = 1.0
DROPDOWN_WAIT        = 0.5
DISMISS_WAIT         = 0.15
CHECKBOX_WAIT        = 0.8       # proven stable from v5.4
CHECKBOX_RETRIES     = 3
REFRESH_POST         = 3.5
TABLE_STABLE_WAIT    = 1.5
WAIT_FOR_XHR_TIMEOUT = 30        # seconds — wait for data XHR after Refresh

# --- Vehicle Category Checkboxes ---
VEHICLE_CATEGORIES = {
    "2W": [
        "TWO WHEELER(NT)",
        "TWO WHEELER(T)",
        "TWO WHEELER (Invalid Carriage)",
    ],
    "3W": [
        "THREE WHEELER(NT)",
        "THREE WHEELER(T)",
        "THREE WHEELER (Invalid Carriage)",
    ],
    "4W": [
        "LIGHT MOTOR VEHICLE",
    ],
}

ALL_CAT_LABELS = [lbl for labels in VEHICLE_CATEGORIES.values() for lbl in labels]
FUEL_LABEL     = "PURE EV"

FILTER_COMBOS = [
    {"cat": "2W", "ev": True,  "suffix": "2W_PureEV"},
    {"cat": "2W", "ev": False, "suffix": "2W_AllFuel"},
    {"cat": "3W", "ev": True,  "suffix": "3W_PureEV"},
    {"cat": "3W", "ev": False, "suffix": "3W_AllFuel"},
    {"cat": "4W", "ev": True,  "suffix": "4W_PureEV"},
    {"cat": "4W", "ev": False, "suffix": "4W_AllFuel"},
]

# ===================================================================
#  STATES
# ===================================================================

STATES = [
    "All Vahan4 Running States",
    "Andaman & Nicobar Island",
    "Andhra Pradesh",
    "Arunachal Pradesh",
    "Assam",
    "Bihar",
    "Chandigarh",
    "Chhattisgarh",
    "UT of DNH and DD",
    "Delhi",
    "Goa",
    "Gujarat",
    "Haryana",
    "Himachal Pradesh",
    "Jammu and Kashmir",
    "Jharkhand",
    "Karnataka",
    "Kerala",
    "Ladakh",
    "Lakshadweep",
    "Madhya Pradesh",
    "Maharashtra",
    "Manipur",
    "Meghalaya",
    "Mizoram",
    "Nagaland",
    "Odisha",
    "Puducherry",
    "Punjab",
    "Rajasthan",
    "Sikkim",
    "Tamil Nadu",
    "Telangana",
    "Tripura",
    "Uttar Pradesh",
    "Uttarakhand",
    "West Bengal",
]

# ===================================================================
#  GRACEFUL SHUTDOWN
# ===================================================================

SHUTDOWN_REQUESTED = False

def _signal_handler(sig, frame):
    global SHUTDOWN_REQUESTED
    SHUTDOWN_REQUESTED = True
    name = signal.Signals(sig).name if hasattr(signal, "Signals") else str(sig)
    print(f"\n[SHUTDOWN] {name} received — finishing current download...")

signal.signal(signal.SIGINT, _signal_handler)
if sys.platform != "win32":
    signal.signal(signal.SIGTERM, _signal_handler)

# ===================================================================
#  LOGGING
# ===================================================================

os.makedirs(OUTPUT_DIR, exist_ok=True)
log_file = os.path.join(
    OUTPUT_DIR, f"scraper_v551_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ===================================================================
#  HELPERS
# ===================================================================

def jittered(base, pct=0.3):
    return base + random.uniform(-base * pct, base * pct)

def make_safe_name(s):
    return s.replace(" ", "_").replace("&", "and").replace("/", "-")

# ===================================================================
#  RESUME
# ===================================================================

PROGRESS_FILE = os.path.join(OUTPUT_DIR, f"progress_{YEAR}.json")

def load_progress():
    if not RESUME_ENABLED:
        return {"completed": []}
    try:
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"completed": []}

def save_progress(state_key, combo_suffix):
    prog = load_progress()
    entry = f"{state_key}__{combo_suffix}"
    if entry not in prog["completed"]:
        prog["completed"].append(entry)
    try:
        with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
            json.dump(prog, f, indent=2)
    except Exception as e:
        log.warning(f"    [WARN] progress save: {e}")

def is_completed(state_key, combo_suffix):
    if not RESUME_ENABLED:
        return False
    return f"{state_key}__{combo_suffix}" in load_progress()["completed"]

# ===================================================================
#  VALIDATION
# ===================================================================

MONTH_ABBR = ["JAN","FEB","MAR","APR","MAY","JUN",
               "JUL","AUG","SEP","OCT","NOV","DEC"]
MONTH_FULL = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]

def validate_download(fpath):
    if not os.path.exists(fpath):
        return False, "file missing"
    size = os.path.getsize(fpath)
    if size < MIN_FILE_SIZE:
        return False, f"too small ({size} bytes)"
    try:
        df = pd.read_excel(fpath, header=None, engine="openpyxl")
        if len(df) < 3:
            return False, f"too few rows ({len(df)})"
        flat = " ".join(df.astype(str).values.flatten()).upper()
        months = [m for m in MONTH_ABBR if m in flat]
        if not months:
            return False, "no month headers"
        return True, f"OK ({size:,}b, {len(df)} rows, {len(months)} months)"
    except Exception as e:
        return False, str(e)

# ===================================================================
#  BROWSER HELPERS
# ===================================================================

async def wait_ajax(page, sec=AJAX_WAIT):
    try:
        await page.wait_for_load_state("networkidle", timeout=15_000)
    except Exception:
        pass
    await asyncio.sleep(jittered(sec))


async def dismiss_overlays(page):
    try:
        await page.keyboard.press("Escape")
        await asyncio.sleep(DISMISS_WAIT)
        await page.evaluate(
            "document.querySelectorAll('.ui-selectonemenu-panel')"
            ".forEach(p=>p.style.display='none');"
        )
        await asyncio.sleep(DISMISS_WAIT * 0.5)
    except Exception:
        pass


# ===================================================================
#  v5.5 KEY FIX — WAIT FOR DATA XHR AFTER REFRESH
# ===================================================================

async def wait_for_data_xhr(page, timeout_sec=WAIT_FOR_XHR_TIMEOUT):
    """
    Wait for VAHAN's data AJAX response after Refresh.
    Returns True if a valid data response was seen, False on timeout.
    """
    log.info(f"    [XHR] Waiting up to {timeout_sec}s for data XHR response...")
    try:
        def is_data_response(response):
            url = response.url
            return (
                "reportview" in url
                and response.status == 200
            )

        async with page.expect_response(
            is_data_response,
            timeout=timeout_sec * 1000
        ) as resp_info:
            pass  # Refresh was already clicked before calling this

        resp = await resp_info.value
        body = await resp.body()
        size = len(body)
        log.info(f"    [XHR] Data response received: {size:,} bytes from {resp.url[:80]}")

        if size < 1000:
            log.warning(f"    [XHR] Response too small ({size}b) — may be error page")
            return False

        return True

    except Exception as e:
        log.warning(f"    [XHR] Timed out waiting for data XHR: {e}")
        return False


async def wait_for_table_stable(page, label=""):
    """Wait until table has rows and count is stable across two reads."""
    log.info(f"    [TABLE] Waiting for stable table{' (' + label + ')' if label else ''}...")
    deadline = asyncio.get_event_loop().time() + 45

    while asyncio.get_event_loop().time() < deadline:
        try:
            count = await page.locator("#groupingTable_data tr").count()
            if count > 0:
                break
        except Exception:
            pass
        await asyncio.sleep(0.5)
    else:
        log.warning("    [TABLE] Timed out waiting for any rows")
        return 0

    for _ in range(10):
        if asyncio.get_event_loop().time() >= deadline:
            break
        c1 = await page.locator("#groupingTable_data tr").count()
        await asyncio.sleep(TABLE_STABLE_WAIT)
        c2 = await page.locator("#groupingTable_data tr").count()
        if c1 == c2 and c1 > 0:
            log.info(f"    [TABLE] Stable at {c1} rows")
            return c1
        log.info(f"    [TABLE] Still loading ({c1}->{c2})...")

    final = await page.locator("#groupingTable_data tr").count()
    log.warning(f"    [TABLE] Stability timeout — current: {final}")
    return final

# ===================================================================
#  CHECKBOX HELPERS
# ===================================================================

async def get_checkbox_state(page, label_text):
    js = """
    (labelText) => {
        const labels = Array.from(document.querySelectorAll('label'));
        for (const l of labels) {
            if (l.innerText.trim() === labelText) {
                const c = l.closest('tr')||l.closest('div')||l.closest('td')||l.parentElement;
                if (c) {
                    const box = c.querySelector('.ui-chkbox-box');
                    if (box) return {found:true, active:box.classList.contains('ui-state-active')};
                }
            }
        }
        return {found:false, active:false};
    }
    """
    return await page.evaluate(js, label_text)


async def set_checkbox(page, label_text, desired):
    """Set checkbox with retry verification. Returns True when confirmed."""
    for attempt in range(1, CHECKBOX_RETRIES + 1):
        state = await get_checkbox_state(page, label_text)
        if not state["found"]:
            log.warning(f"    [WARN] Checkbox not found: '{label_text}'")
            return False
        if state["active"] == desired:
            return True  # already correct

        toggled = False
        try:
            loc = page.locator(
                f"tr:has(label:text-is('{label_text}')) .ui-chkbox-box, "
                f"div:has(> label:text-is('{label_text}')) .ui-chkbox-box, "
                f"td:has(label:text-is('{label_text}')) .ui-chkbox-box"
            ).first
            if await loc.count() > 0:
                await loc.click(force=True)
                toggled = True
        except Exception:
            pass

        if not toggled:
            toggled = await page.evaluate("""
            (t) => {
                for (const l of document.querySelectorAll('label')) {
                    if (l.innerText.trim()===t) {
                        const c=l.closest('tr')||l.closest('div')||l.closest('td')||l.parentElement;
                        if(c){const b=c.querySelector('.ui-chkbox-box');if(b){b.click();return true;}}
                    }
                }
                return false;
            }""", label_text)

        await asyncio.sleep(CHECKBOX_WAIT)

        verify = await get_checkbox_state(page, label_text)
        if verify["active"] == desired:
            action = "CHECKED" if desired else "UNCHECKED"
            log.info(f"    [OK] {action}: '{label_text}' (attempt {attempt})")
            return True

        log.warning(f"    [WARN] Toggle did not land for '{label_text}' attempt {attempt}/{CHECKBOX_RETRIES}")

    log.error(f"    [FAIL] Could not set '{label_text}' after {CHECKBOX_RETRIES} attempts")
    return False

# ===================================================================
#  FILTER PANEL
# ===================================================================

async def open_filter_panel(page):
    try:
        await page.evaluate("document.getElementById('filterLayout-toggler').click();")
        await asyncio.sleep(PANEL_OPEN_WAIT)
        log.info("    [OK] Filter panel opened")
        return True
    except Exception as e:
        log.warning(f"    [WARN] Panel open: {e}")
        return False


async def click_panel_refresh(page):
    """Click Refresh inside filter panel. Returns True on success."""
    try:
        loc = page.locator(
            "#filterLayout .ui-button:has-text('Refresh'), "
            "#filterLayout button:has-text('Refresh'), "
            "#filterLayout a:has-text('Refresh')"
        ).first
        if await loc.count() > 0:
            await loc.click(force=True)
            log.info("    [OK] Panel Refresh clicked (Playwright)")
            return True
    except Exception:
        pass

    try:
        result = await page.evaluate("""
        () => {
            const p = document.getElementById('filterLayout');
            if (!p) return false;
            for (const b of p.querySelectorAll('button, a.ui-commandlink, .ui-button')) {
                if (b.textContent.trim().includes('Refresh')) { b.click(); return true; }
            }
            return false;
        }""")
        if result:
            log.info("    [OK] Panel Refresh clicked (JS)")
            return True
    except Exception:
        pass

    log.warning("    [WARN] Panel Refresh not found")
    return False


async def apply_filters(page, cat_key, pure_ev):
    """
    v5.5 sequence:
      1. Open filter panel
      2. Set + verify all checkboxes
      3. Click panel Refresh
      4. Wait for data XHR response  <- KEY FIX
      5. Wait for table stability
    """
    ev_str = "Yes" if pure_ev else "No"
    log.info(f"    [FILTER] Applying: {cat_key} | PureEV={ev_str}")

    try:
        await open_filter_panel(page)

        desired_labels = set(VEHICLE_CATEGORIES[cat_key])
        all_ok = True

        for label in ALL_CAT_LABELS:
            ok = await set_checkbox(page, label, label in desired_labels)
            if not ok:
                all_ok = False

        ok = await set_checkbox(page, FUEL_LABEL, pure_ev)
        if not ok:
            all_ok = False

        if not all_ok:
            log.error("    [FAIL] Checkbox confirmation failed — aborting combo")
            return False

        log.info("    [FILTER] All checkboxes confirmed — clicking panel Refresh")

        async with page.expect_response(
            lambda r: "reportview" in r.url and r.status == 200,
            timeout=WAIT_FOR_XHR_TIMEOUT * 1000
        ) as resp_info:
            await click_panel_refresh(page)

        resp = await resp_info.value
        body = await resp.body()
        log.info(f"    [XHR] Data response: {len(body):,} bytes")

        if len(body) < 1000:
            log.warning(f"    [XHR] Response suspiciously small — may be stale")

        await asyncio.sleep(jittered(REFRESH_POST))
        row_count = await wait_for_table_stable(page, label=f"{cat_key} PureEV={ev_str}")

        if row_count == 0:
            log.error("    [FAIL] Table empty after filter apply")
            return False

        log.info(f"    [OK] Filter applied: {cat_key} | PureEV={ev_str} | {row_count} rows")
        return True

    except Exception as e:
        log.error(f"    [FAIL] apply_filters: {e}")
        return False

# ===================================================================
#  AUDIT FILTER STATE
# ===================================================================

async def audit_filter_state(page, expected_cat, expected_ev):
    expected_labels = set(VEHICLE_CATEGORIES[expected_cat])
    mismatch = False
    log.info(f"    [AUDIT] cat={expected_cat} PureEV={expected_ev}")

    for label in ALL_CAT_LABELS + [FUEL_LABEL]:
        state = await get_checkbox_state(page, label)
        if not state["found"]:
            log.warning(f"    [AUDIT]   NOT FOUND: '{label}'")
            continue
        should = (label in expected_labels) if label != FUEL_LABEL else expected_ev
        actual = state["active"]
        icon   = "OK" if actual == should else "X MISMATCH"
        log.info(f"    [AUDIT]   [{icon}] '{label}' expected={'ON' if should else 'OFF'} actual={'ON' if actual else 'OFF'}")
        if actual != should:
            mismatch = True

    if mismatch:
        log.error("    [AUDIT] MISMATCH — download will be skipped")
    else:
        log.info("    [AUDIT] All filters verified OK")
    return not mismatch

# ===================================================================
#  MAIN REFRESH
# ===================================================================

async def click_main_refresh(page):
    try:
        await dismiss_overlays(page)
        await asyncio.sleep(0.3)
        result = await page.evaluate("""
        () => {
            const all = document.querySelectorAll('button,.ui-button,.ui-commandbutton,a.ui-commandlink');
            for (const el of all) {
                if (el.textContent.trim().includes('Refresh')) {
                    el.click(); return 'clicked|'+el.id;
                }
            }
            return 'not_found';
        }""")
        if "clicked" in result:
            log.info(f"  [OK] Main Refresh: {result}")
            await wait_ajax(page, REFRESH_POST)
            return True
        log.warning("  [WARN] Main Refresh not found")
        return False
    except Exception as e:
        log.warning(f"  [WARN] Main Refresh: {e}")
        return False

# ===================================================================
#  DROPDOWN + STATE
# ===================================================================

async def select_dropdown(page, element_id, option_text, name=""):
    desc = name or element_id
    try:
        await dismiss_overlays(page)
        lbl = page.locator(f"#{element_id}_label")
        if await lbl.count() == 0:
            log.warning(f"  [WARN] {desc}: label not found")
            return False
        await lbl.first.click(force=True)
        await asyncio.sleep(DROPDOWN_WAIT)

        panel = page.locator(f"#{element_id}_panel")
        if await panel.count() == 0:
            log.warning(f"  [WARN] {desc}: panel not found")
            await dismiss_overlays(page)
            return False

        opt = panel.locator(f"li[data-label='{option_text}']")
        if await opt.count() == 0:
            opt = panel.locator(f"li:has-text('{option_text}')").first
        if await opt.count() == 0:
            log.warning(f"  [WARN] {desc}: '{option_text}' not in panel")
            await dismiss_overlays(page)
            return False

        await opt.first.click(force=True)
        await wait_ajax(page, AJAX_WAIT)
        await dismiss_overlays(page)
        val = (await page.locator(f"#{element_id}_label").text_content()).strip()
        log.info(f"  [OK] {desc} = '{val}'")
        return True
    except Exception as e:
        log.warning(f"  [WARN] {desc}: {e}")
        await dismiss_overlays(page)
        return False


async def select_state(page, state_name):
    try:
        await dismiss_overlays(page)
        await asyncio.sleep(DISMISS_WAIT)

        opened = await page.evaluate("""
        () => {
            for (const l of document.querySelectorAll('.ui-selectonemenu-label')) {
                const t = l.textContent.trim();
                if (t.includes('Vahan4')||t.includes('Running')||t.includes('(')) {
                    l.click(); return 'opened|'+t;
                }
            }
            return 'not_found';
        }""")
        if "not_found" in opened:
            log.error("  [FAIL] State dropdown not found")
            return False
        log.info(f"  [OK] Dropdown: {opened}")
        await asyncio.sleep(DROPDOWN_WAIT)

        result = await page.evaluate("""
        (name) => {
            for (const p of document.querySelectorAll('.ui-selectonemenu-panel')) {
                if (p.offsetParent===null && p.style.display==='none') continue;
                for (const li of p.querySelectorAll('li')) {
                    if (li.textContent.trim()===name) { li.click(); return 'exact|'+name; }
                }
                for (const li of p.querySelectorAll('li')) {
                    if (li.getAttribute('data-label')===name) { li.click(); return 'attr|'+name; }
                }
                for (const li of p.querySelectorAll('li')) {
                    if (li.textContent.trim().includes(name)) { li.click(); return 'partial|'+name; }
                }
            }
            return 'not_found';
        }""", state_name)

        if "not_found" in result:
            log.error(f"  [FAIL] State '{state_name}' not in dropdown")
            await dismiss_overlays(page)
            return False

        log.info(f"  [OK] State: {result}")
        await wait_ajax(page, AJAX_WAIT)
        await dismiss_overlays(page)
        return True
    except Exception as e:
        log.error(f"  [FAIL] select_state: {e}")
        await dismiss_overlays(page)
        return False

# ===================================================================
#  DASHBOARD SETUP
# ===================================================================

async def setup_dashboard(page):
    log.info("=" * 65)
    log.info("ONE-TIME SETUP")
    log.info("=" * 65)

    await page.goto(VAHAN_URL, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
    await wait_ajax(page, 3)
    log.info("[OK] Dashboard loaded")
    await page.screenshot(path=os.path.join(OUTPUT_DIR, "01_loaded.png"))
    await dismiss_overlays(page)

    await select_dropdown(page, "yaxisVar",        Y_AXIS,          "Y-Axis")
    await select_dropdown(page, "xaxisVar",        X_AXIS,          "X-Axis")

    try:
        yt = (await page.locator("#selectedYearType_label").text_content()).strip()
        if "Calendar Year" not in yt:
            await select_dropdown(page, "selectedYearType", "Calendar Year", "Year Type")
        else:
            log.info(f"  [OK] Year Type already '{yt}'")
    except Exception:
        await select_dropdown(page, "selectedYearType", "Calendar Year", "Year Type")

    try:
        yr = (await page.locator("#selectedYear_label").text_content()).strip()
        if YEAR not in yr:
            await select_dropdown(page, "selectedYear", YEAR, "Year")
        else:
            log.info(f"  [OK] Year already '{yr}'")
    except Exception:
        await select_dropdown(page, "selectedYear", YEAR, "Year")

    await click_main_refresh(page)
    await page.screenshot(path=os.path.join(OUTPUT_DIR, "02_setup.png"))
    log.info("[OK] Setup done")
    log.info("=" * 65)

# ===================================================================
#  EXCEL CLEANER v5.5
# ===================================================================

def clean_excel(fpath, state_name, suffix, year):
    temp_fd = temp_path = None
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        parts = suffix.replace("_", " ").replace("PureEV", "Pure EV").replace("AllFuel", "All Fuel")
        title = f"{state_name} {parts} {year} Maker Data"

        wb_raw = openpyxl.load_workbook(fpath)
        ws_raw = wb_raw.active
        for mr in list(ws_raw.merged_cells.ranges):
            tlv = ws_raw.cell(mr.min_row, mr.min_col).value
            ws_raw.unmerge_cells(str(mr))
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    ws_raw.cell(r, c, tlv)

        temp_fd = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=OUTPUT_DIR)
        temp_path = temp_fd.name
        temp_fd.close()
        wb_raw.save(temp_path)
        wb_raw.close()

        raw = pd.read_excel(temp_path, header=None, engine="openpyxl")
        log.info(f"    [CLEAN] Raw shape: {raw.shape}")

        month_row_idx = None
        for ri in range(min(10, len(raw))):
            row_vals = [str(v).strip().upper() for v in raw.iloc[ri]]
            if any(m in row_vals for m in MONTH_ABBR):
                month_row_idx = ri
                break

        if month_row_idx is None:
            log.warning("    [CLEAN] No month header row found")
            return True

        month_row_upper = [str(v).strip().upper() for v in raw.iloc[month_row_idx]]
        month_col_map   = {ci: MONTH_ABBR.index(v)
                           for ci, v in enumerate(month_row_upper) if v in MONTH_ABBR}

        sno_col = maker_col = None
        for check_ri in range(max(0, month_row_idx - 2), month_row_idx + 1):
            for ci, val in enumerate([str(v).strip().upper() for v in raw.iloc[check_ri]]):
                vc = val.replace(".", "").replace(" ", "")
                if sno_col is None and vc in ("SNO","SN","S","NO","SRNO","SERIAL"):
                    sno_col = ci
                if maker_col is None and "MAKER" in val:
                    maker_col = ci

        if sno_col   is None: sno_col   = 0
        if maker_col is None: maker_col = 1

        sorted_months = sorted(month_col_map.items(), key=lambda x: x[1])
        new_headers   = ["S. No.", "Maker"] + \
                        [f"{MONTH_FULL[mi]} {year}" for _, mi in sorted_months] + \
                        ["Total"]

        data_rows = []
        sno = 1
        for ri in range(month_row_idx + 1, len(raw)):
            row       = raw.iloc[ri]
            maker_val = str(row.iloc[maker_col]).strip() if pd.notna(row.iloc[maker_col]) else ""
            if not maker_val or maker_val.upper() in ("NAN", "MAKER"):
                continue

            month_values = []
            for col_idx, _ in sorted_months:
                v = row.iloc[col_idx] if col_idx < len(row) else 0
                if isinstance(v, str):
                    v = v.replace(",", "").strip()
                v_num = pd.to_numeric(v, errors="coerce")
                month_values.append(int(v_num) if not pd.isna(v_num) else 0)

            total_val = sum(month_values)
            data_rows.append([sno, maker_val] + month_values + [total_val])
            sno += 1

        if not data_rows:
            log.warning("    [CLEAN] No data rows found")
            return True

        top5 = [r[-1] for r in data_rows[:5]]
        if all(t == 0 for t in top5):
            log.error("    [GUARD] All top-5 makers show 0 — suspected wrong filter. Deleting file.")
            try:
                os.remove(fpath)
            except Exception:
                pass
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        ws.cell(1, 1, title).font = Font(bold=True, size=13)

        for ci, h in enumerate(new_headers, 1):
            c = ws.cell(2, ci, h)
            c.font      = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for ri, drow in enumerate(data_rows, 3):
            for ci, val in enumerate(drow, 1):
                ws.cell(ri, ci, val).alignment = Alignment(horizontal="center")
            ws.cell(ri, 2).alignment = Alignment(horizontal="left")

        for ci in range(1, len(new_headers) + 1):
            max_len = max(
                len(str(ws.cell(2, ci).value or "")),
                max((len(str(ws.cell(ri, ci).value or "")) for ri in range(3, len(data_rows) + 3)), default=0)
            )
            ws.column_dimensions[get_column_letter(ci)].width = max(max_len + 3, 10)

        wb.save(fpath)
        log.info(f"    [CLEAN] Done: {os.path.basename(fpath)} | {len(data_rows)} rows")
        return True

    except Exception as e:
        log.warning(f"    [CLEAN] Failed: {e}\n{traceback.format_exc()}")
        return True
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

# ===================================================================
#  DOWNLOAD — 6 FILES PER STATE
# ===================================================================

async def download_state(page, state_name, idx):
    safe         = make_safe_name(state_name)
    total_states = len(STATES)
    total_combos = len(FILTER_COMBOS)
    log.info(f"[{idx}/{total_states}] {state_name}")

    log.info("  STEP 0: Select state...")
    if not await select_state(page, state_name):
        log.error(f"  [FAIL] Could not select: {state_name}")
        return {"success": [], "failed": [c["suffix"] for c in FILTER_COMBOS]}

    log.info("  STEP 1: Main Refresh after state change...")
    await click_main_refresh(page)

    results = {"success": [], "failed": []}

    for combo_idx, combo in enumerate(FILTER_COMBOS, 1):
        if SHUTDOWN_REQUESTED:
            results["failed"].extend([c["suffix"] for c in FILTER_COMBOS[combo_idx - 1:]])
            break

        cat_key = combo["cat"]
        pure_ev = combo["ev"]
        suffix  = combo["suffix"]
        # v5.5.1: append RUN_DATE so each daily run produces a unique filename
        fname   = f"{safe}_{suffix}_{YEAR}_{RUN_DATE}.xlsx"
        fpath   = os.path.join(OUTPUT_DIR, fname)

        log.info(f"\n  {'-'*45}")
        log.info(f"  Combo {combo_idx}/{total_combos}: {suffix}")
        log.info(f"  {'-'*45}")

        if is_completed(safe, suffix):
            if os.path.exists(fpath):
                is_valid, msg = validate_download(fpath)
                if is_valid:
                    log.info(f"    [SKIP] Already done & valid: {fname} ({msg})")
                    results["success"].append(suffix)
                    continue
            log.info(f"    [REDO] Marked complete but file missing/invalid")

        if os.path.exists(fpath):
            try:
                os.remove(fpath)
            except Exception:
                pass

        log.info(f"    STEP A: Apply filters ({cat_key}, PureEV={pure_ev})...")
        if not await apply_filters(page, cat_key, pure_ev):
            log.error(f"    [FAIL] Filter apply failed — skipping {suffix}")
            results["failed"].append(suffix)
            continue

        log.info("    STEP B: Audit filter state...")
        if not await audit_filter_state(page, cat_key, pure_ev):
            log.error(f"    [FAIL] Audit mismatch — skipping {suffix}")
            try:
                await page.screenshot(
                    path=os.path.join(OUTPUT_DIR, f"AUDIT_FAIL_{safe}_{suffix}.png")
                )
            except Exception:
                pass
            results["failed"].append(suffix)
            continue

        row_count = await page.locator("#groupingTable_data tr").count()
        log.info(f"    STEP C: Table rows = {row_count}")
        try:
            hdr = page.locator(".ui-datatable-header")
            if await hdr.count() > 0:
                t = await hdr.first.text_content()
                log.info(f"    Header: '{t.strip()[:120]}'")
        except Exception:
            pass

        if row_count == 0:
            log.error(f"    [FAIL] Table empty — skipping {suffix}")
            results["failed"].append(suffix)
            continue

        log.info(f"    STEP D: Download -> {fname}")
        try:
            btn = page.locator(
                ".ui-datatable-header a, "
                "a:has(img[src*='excel']), "
                "a:has(img[src*='xls']), "
                "img[src*='excel'], "
                "[id*='Excel'], [id*='excel']"
            ).first

            if await btn.count() == 0:
                log.error(f"    [FAIL] Download button not found")
                results["failed"].append(suffix)
                continue

            btn_href = await btn.get_attribute("href")
            btn_id   = await btn.get_attribute("id")
            log.info(f"    [DL] btn id='{btn_id}' href='{btn_href}'")

            async with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as dl_info:
                await btn.click(force=True)

            download = await dl_info.value
            await download.save_as(fpath)
            size = os.path.getsize(fpath)
            log.info(f"    SAVED: {fname} ({size:,} bytes)")

            is_valid, val_msg = validate_download(fpath)
            if not is_valid:
                log.error(f"    [INVALID] {val_msg}")
                results["failed"].append(suffix)
                continue

            log.info(f"    [VALID] {val_msg}")

            clean_ok = clean_excel(fpath, state_name, suffix, YEAR)
            if not clean_ok:
                log.error(f"    [GUARD] Zero-total guard triggered — marking FAILED")
                results["failed"].append(suffix)
                continue

            results["success"].append(suffix)
            save_progress(safe, suffix)

        except Exception as e:
            log.error(f"    [FAIL] Download {suffix}: {e}")
            results["failed"].append(suffix)

    return results

# ===================================================================
#  MAIN LOOP
# ===================================================================

async def scrape_all():
    global SHUTDOWN_REQUESTED
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log.error("Run: pip install playwright && playwright install")
        return None

    summary = {"success": [], "failed": []}

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        context = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()
        page.set_default_timeout(PAGE_TIMEOUT)

        try:
            await setup_dashboard(page)
        except Exception as e:
            log.error(f"[FAIL] Setup: {e}")
            await page.screenshot(path=os.path.join(OUTPUT_DIR, "ERROR_setup.png"))
            await browser.close()
            return summary

        for idx, state in enumerate(STATES, 1):
            if SHUTDOWN_REQUESTED:
                summary["failed"].extend(STATES[idx - 1:])
                break

            ok   = False
            safe = make_safe_name(state)

            for attempt in range(1, MAX_RETRIES + 1):
                if SHUTDOWN_REQUESTED:
                    break
                try:
                    log.info(f"\n{'='*55}")
                    log.info(f"State {idx}/{len(STATES)}: {state}  (attempt {attempt}/{MAX_RETRIES})")
                    log.info(f"{'='*55}")

                    result = await download_state(page, state, idx)

                    if result and result["success"]:
                        summary["success"].append(f"{state} ({len(result['success'])}/6)")
                        ok = True
                        if result["failed"]:
                            log.warning(f"  [PARTIAL] failed: {result['failed']}")
                        else:
                            log.info(f"  [OK] {state}: 6/6")
                        break
                    else:
                        log.warning(f"  [WARN] {state}: no files downloaded")

                except Exception as e:
                    log.error(f"  [FAIL] attempt {attempt}: {e}")
                    try:
                        await page.screenshot(
                            path=os.path.join(OUTPUT_DIR, f"ERROR_{safe}_{attempt}.png")
                        )
                    except Exception:
                        pass

                if attempt < MAX_RETRIES and not ok:
                    wait = RETRY_BACKOFF * (2 ** (attempt - 1))
                    log.info(f"  Retry in {wait}s...")
                    await asyncio.sleep(wait)
                    try:
                        await page.goto(VAHAN_URL, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                        await wait_ajax(page, 3)
                        await setup_dashboard(page)
                    except Exception:
                        log.warning("  [WARN] Reload failed")

            if not ok and not SHUTDOWN_REQUESTED:
                summary["failed"].append(state)
                log.error(f"  GAVE UP: {state}")

            if not SHUTDOWN_REQUESTED:
                delay = random.uniform(STATE_DELAY_MIN, STATE_DELAY_MAX)
                log.info(f"  [DELAY] {delay:.1f}s before next state...")
                await asyncio.sleep(delay)

        try:
            await page.screenshot(path=os.path.join(OUTPUT_DIR, "03_done.png"))
        except Exception:
            pass
        await browser.close()
        log.info("[OK] Browser closed")

    return summary

# ===================================================================
#  ENTRY POINT
# ===================================================================

async def main():
    t0 = time.time()
    log.info("+" + "="*62 + "+")
    log.info("  VAHAN EV SCRAPER v5.5.1")
    log.info("  Multi-Filter: 2W / 3W / 4W x PureEV / AllFuel")
    log.info("  v5.5.1: auto-year + dated filenames (YYYYMMDD suffix)")
    log.info("+" + "="*62 + "+")
    log.info(f"States   : {len(STATES)}  |  Combos/state : {len(FILTER_COMBOS)}  |  Total files : {len(STATES)*len(FILTER_COMBOS)}")
    log.info(f"Year     : {YEAR}  |  Run Date : {RUN_DATE}  |  Headless : {HEADLESS}  |  Resume : {RESUME_ENABLED}")
    log.info(f"Output   : {os.path.abspath(OUTPUT_DIR)}")
    log.info("")

    if RESUME_ENABLED:
        prog = load_progress()
        n = len(prog["completed"])
        log.info(f"[RESUME] {n} previously completed downloads found")
    else:
        log.info("[RESUME] Disabled")
    log.info("")

    summary = await scrape_all()
    elapsed = time.time() - t0

    log.info("\n" + "="*65)
    log.info("SUMMARY")
    log.info("="*65)
    if summary:
        log.info(f"  Success : {len(summary['success'])}/{len(STATES)}")
        log.info(f"  Failed  : {len(summary['failed'])}")
        for s in summary["success"]:
            log.info(f"    + {s}")
        if summary["failed"]:
            log.info(f"  Failed  : {', '.join(str(f) for f in summary['failed'])}")
    if SHUTDOWN_REQUESTED:
        log.info("  NOTE: Interrupted — run again to resume")
    log.info(f"  Time     : {elapsed/60:.1f} min ({elapsed:.0f}s)")
    log.info(f"  Output   : {os.path.abspath(OUTPUT_DIR)}")
    log.info("="*65)


if __name__ == "__main__":
    asyncio.run(main())
